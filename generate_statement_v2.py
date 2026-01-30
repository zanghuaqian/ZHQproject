# -*- coding: utf-8 -*-
"""
盛意旺采购商城对账单生成脚本
根据原数据表格和交易对账单文件生成标准格式的对账单

使用方法:
    python scripts/generate_statement.py <原数据文件路径> <交易对账单文件路径> <月份> [输出路径]
    
示例:
    python scripts/generate_statement.py data.xlsx reference.xlsx 1
    python scripts/generate_statement.py data.xlsx reference.xlsx 1 output.xlsx
"""

import sys
import os
from datetime import datetime
import warnings

try:
    import pandas as pd
except ImportError:
    print("错误: 未安装pandas")
    print("请运行: pip install pandas")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
    from openpyxl.styles.numbers import BUILTIN_FORMATS
    warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
except ImportError:
    print("错误: 未安装openpyxl")
    print("请运行: pip install openpyxl")
    sys.exit(1)

# 对账单标准列（35列）
STATEMENT_COLUMNS = [
    '月份', '下订单部门', '产品订货单号', '商户号', '商户名称', '省市',
    '机具厂商', '型号', '发货数量（台）', '单价', '优惠价', '应收金额',
    '实收金额', '收款时间', '实际发货台数', '实际发货金额', '快递费',
    '收款金额', '退款-退快递费', '退款-退货', '退货-台数', '钱包款（实际收入金额）',
    '线下款', '线上支付编号', '发票种类', '发票号码', '支付渠道',
    '商城支付订单号', '退款单号', '退款时间', '二次退款', '二次退款时间',
    'MSP单号', '（预留列）', '备注'
]

# 下单部门映射规则
DEPARTMENT_MAPPING = {
    '产品运营中心': ['徐晶', '尤晨伟', '陆瑶瑶'],
    '生态合作部': ['冯勇', '史卫锋', '毛凌霄', '马勇', '贾森', '戴志坤', '李晨辉', 
                  '周黎琼', '王政', '耿立阳', '宋帅康', '曾东光', '熊三亮', 
                  '兰其发', '刘晓辉', '于杰', '陈飞'],
    '金融合作中心': ['董国伟', '康惠仁', '葛嘉成', '邱从伟', '王遥遥', '倪斌', '吴俊', 
                  '陆文斌', '郭米加', '赵振鹏', '孙厚强', '葛宏洲', '韩方明月', 
                  '汪亮', '庄思佳', '刘欣雨'],
    '资金赋能中心': ['夏俊', '李超', '蒋尚斌', '胡孝儒', '常赛', '汪婷婷']
}

# 列宽设置
COLUMN_WIDTHS = {
    'A': 10, 'B': 20, 'C': 40, 'D': 12, 'E': 25, 'F': 20, 'G': 20, 'H': 30,
    'I': 12, 'J': 12, 'K': 12, 'L': 12, 'M': 12, 'N': 12, 'O': 12, 'P': 12,
    'Q': 12, 'R': 12, 'S': 15, 'T': 12, 'U': 12, 'V': 18, 'W': 12, 'X': 35,
    'Y': 12, 'Z': 20, 'AA': 12, 'AB': 40, 'AC': 35, 'AD': 12, 'AE': 12,
    'AF': 12, 'AG': 25, 'AH': 15, 'AI': 30
}

def find_column(df, possible_names):
    """查找可能的列名"""
    if df is None:
        return None
    for name in possible_names:
        if name in df.columns:
            return name
    return None

def map_merchant_bd_to_department(merchant_bd):
    """根据商户BD映射到下单部门"""
    if pd.isna(merchant_bd):
        return None
    merchant_bd = str(merchant_bd).strip()
    for department, names in DEPARTMENT_MAPPING.items():
        if merchant_bd in names:
            return department
    return None

def process_supplier(supplier):
    """处理发货方字段"""
    if pd.isna(supplier):
        return None
    supplier = str(supplier).strip()
    if supplier.lower() == 'system':
        return '盛意旺自营'
    elif supplier.startswith('厂商-'):
        return supplier.replace('厂商-', '', 1)
    else:
        return supplier

def map_invoice_type(need_invoice):
    """映射发票种类"""
    if pd.isna(need_invoice):
        return '不开票'
    need_invoice = str(need_invoice).strip().lower()
    if need_invoice in ['是', 'yes', 'true', '1', '需要', '开票']:
        return '开票'
    else:
        return '不开票'

def map_payment_channel(merchant_order_no, reference_df):
    """映射支付渠道
    从交易对账单文件的"交易明细表格"工作表中获取
    根据商城支付订单号查询交易明细表中的商户订单号所在订单，获取"支付渠道"字段的值
    """
    if pd.isna(merchant_order_no) or reference_df is None:
        return None
    
    merchant_order_col = find_column(reference_df, ['商户订单号', '订单号', '商城订单号', '商城支付订单号'])
    if not merchant_order_col:
        return None
    
    merchant_order_no_str = str(merchant_order_no).strip()
    reference_df_matched = reference_df.copy()
    reference_df_matched[merchant_order_col] = reference_df_matched[merchant_order_col].astype(str).str.strip()
    
    match = reference_df_matched[reference_df_matched[merchant_order_col] == merchant_order_no_str]
    
    if not match.empty:
        channel_col = find_column(reference_df, ['支付渠道', '支付方式', '渠道'])
        if channel_col:
            for row_idx, matched_row in match.iterrows():
                original_idx = matched_row.name
                channel = reference_df.loc[original_idx, channel_col] if original_idx in reference_df.index else matched_row.get(channel_col, None)
                
                if pd.notna(channel) and str(channel).strip():
                    channel_str = str(channel).strip()
                    channel_upper = channel_str.upper()
                    if '微信' in channel_str or 'WX' in channel_upper or 'WEIXIN' in channel_upper:
                        return '微信'
                    elif '支付宝' in channel_str or 'ALIPAY' in channel_upper or 'ZFB' in channel_upper or 'ALI' in channel_upper:
                        return '支付宝'
                    return channel_str
    return None

def process_refund_info(row, payment_status):
    """处理退款信息"""
    refund_express_fee = None
    refund_amount = None
    refund_quantity = None
    needs_manual_check = False
    
    if payment_status in ['退款成功', 'REFUND_PART_SUCCESS']:
        refund_total = row.get('退款金额', None)
        express_fee = row.get('快递费', None)
        payment_amount = row.get('支付金额', None)
        actual_payment = row.get('实付款', None)
        purchase_quantity = row.get('采购数量', None)
        
        try:
            refund_total = float(refund_total) if pd.notna(refund_total) else None
            express_fee = float(express_fee) if pd.notna(express_fee) else 0.00
            payment_amount = float(payment_amount) if pd.notna(payment_amount) else None
            actual_payment = float(actual_payment) if pd.notna(actual_payment) else None
            purchase_quantity = int(purchase_quantity) if pd.notna(purchase_quantity) else 0
        except (ValueError, TypeError):
            refund_total = None
            express_fee = 0.00
            payment_amount = None
            actual_payment = None
            purchase_quantity = 0
        
        if refund_total is not None:
            if abs(refund_total - express_fee) < 0.01:
                refund_express_fee = refund_total
                refund_amount = 0.00
                refund_quantity = 0
            elif payment_amount is not None and abs(refund_total - payment_amount) < 0.01:
                refund_express_fee = express_fee
                refund_amount = actual_payment if actual_payment is not None else 0.00
                refund_quantity = purchase_quantity
            else:
                needs_manual_check = True
                refund_express_fee = express_fee
                refund_amount = actual_payment if actual_payment is not None else 0.00
                refund_quantity = purchase_quantity
    
    return {
        '退款-退快递费': refund_express_fee,
        '退款-退货': refund_amount,
        '退货-台数': refund_quantity,
        '需要人工校验': needs_manual_check
    }

def find_cross_month_order_in_source(original_order_no, df_source):
    """在原数据表中查找跨月退款对应的原始订单
    
    根据商城支付订单号进行关联匹配（优先级最高）
    
    Args:
        original_order_no: 退款订单原商户订单号（即商城支付订单号）
        df_source: 原数据DataFrame
    
    Returns:
        匹配的订单记录，如果未找到返回None
    """
    if df_source is None or original_order_no is None:
        return None
    
    original_order_str = str(original_order_no).strip()
    
    # 优先通过商城支付单号匹配（PA开头的订单号）
    payment_col = find_column(df_source, ['商城支付单号', '商城支付订单号', 'PA订单号', '支付订单号'])
    if payment_col:
        df_matched = df_source.copy()
        df_matched[payment_col] = df_matched[payment_col].astype(str).str.strip()
        match = df_matched[df_matched[payment_col] == original_order_str]
        if not match.empty:
            print(f"  ✓ 通过商城支付订单号匹配成功: {original_order_str}")
            return match.iloc[0]
    
    # 其次通过订单号匹配（PO开头的订单号）
    order_col = find_column(df_source, ['订单号', '产品订货单号', '商城订单号'])
    if order_col:
        df_matched = df_source.copy()
        df_matched[order_col] = df_matched[order_col].astype(str).str.strip()
        match = df_matched[df_matched[order_col] == original_order_str]
        if not match.empty:
            print(f"  ✓ 通过产品订单号匹配成功: {original_order_str}")
            return match.iloc[0]
    
    print(f"  ✗ 未找到匹配订单: {original_order_str}")
    return None

def generate_statement(source_file, reference_file, month, output_path=None, cross_month_source_files=None):
    """生成对账单
    
    Args:
        source_file: 原数据表格文件路径
        reference_file: 交易对账单文件路径
        month: 对账月份（1-12）
        output_path: 输出文件路径（可选）
        cross_month_source_files: 跨月退款对应月份的原数据文件字典 {year_month: file_path}
    
    Returns:
        输出文件路径
    """
    # 读取原数据表格
    print(f"\n📂 读取原数据文件: {source_file}")
    df_source = pd.read_excel(source_file)
    print(f"   ✓ 共读取 {len(df_source)} 条记录")
    print(f"   ✓ 列名: {list(df_source.columns)[:10]}{'...' if len(df_source.columns) > 10 else ''}")
    
    # 跨月退款对应月份的原数据字典
    cross_month_dfs = {}
    if cross_month_source_files:
        print(f"\n📂 加载跨月退款原数据文件...")
        print(f"   说明: 将根据商城支付订单号进行关联匹配")
        for year_month, file_path in cross_month_source_files.items():
            try:
                cross_month_dfs[year_month] = pd.read_excel(file_path)
                print(f"   ✓ {year_month} 月份: {file_path} ({len(cross_month_dfs[year_month])} 条记录)")
            except Exception as e:
                print(f"   ✗ {year_month} 月份加载失败: {e}")
    
    # 读取交易对账单（从"交易明细表格"工作表，第4行作为表头）
    print(f"\n📂 读取交易对账单: {reference_file}")
    excel_file = pd.ExcelFile(reference_file)
    sheet_names = excel_file.sheet_names
    print(f"   ✓ 工作表: {sheet_names}")
    
    target_sheet = None
    for sheet in sheet_names:
        if '交易明细' in sheet or '明细' in sheet:
            target_sheet = sheet
            break
    
    if target_sheet:
        df_reference = pd.read_excel(reference_file, sheet_name=target_sheet, header=3)
        print(f"   ✓ 使用工作表: {target_sheet}")
    else:
        df_reference = pd.read_excel(reference_file, sheet_name=0, header=3)
        print(f"   ✓ 使用默认工作表: {sheet_names[0]}")
    
    print(f"   ✓ 共读取 {len(df_reference)} 条交易记录")
    
    # 当前年月（用于判断跨月退款）
    current_year = datetime.now().year
    current_month_str = f"{current_year}{month:02d}"
    print(f"\n📅 对账月份: {current_year}年{month}月 ({current_month_str})")
    
    # 过滤订单状态
    print(f"\n🔍 过滤订单状态...")
    if '支付状态' in df_source.columns:
        print(f"   - 原数据包含'支付状态'字段")
        print(f"   - 支付状态分布: {df_source['支付状态'].value_counts().to_dict()}")
        valid_statuses = ['支付成功', '退款成功', 'REFUND_PART_SUCCESS']
        df_filtered = df_source[df_source['支付状态'].isin(valid_statuses)].copy()
        print(f"   ✓ 过滤后剩余 {len(df_filtered)} 条记录 (状态: {valid_statuses})")
    else:
        print(f"   ⚠ 原数据不包含'支付状态'字段，使用全部数据")
        df_filtered = df_source.copy()
        print(f"   ✓ 共 {len(df_filtered)} 条记录")
    
    # 映射字段
    print(f"\n📝 开始映射字段...")
    result_data = []
    cross_month_refund_data = []  # 存储跨月退款订单
    month_value = int(current_month_str)
    
    for idx, row in df_filtered.iterrows():
        # 提取省市信息（从收货人地址中提取前两级）
        province_city = None
        address = row.get('收货人地址', None)
        if pd.notna(address):
            address_str = str(address)
            # 尝试提取省市（例如：上海市上海市 → 上海市上海市，广东省深圳市 → 广东省深圳市）
            import re
            match = re.match(r'(.{2,3}[省市])(.{2,3}[市区县])', address_str)
            if match:
                province_city = match.group(1) + match.group(2)
        
        record = {
            '月份': month_value,
            '商户号': row.get('买家ID', None),
            '商户名称': row.get('买家名称', None),
            '产品订货单号': row.get('订单号', None),
            '机具厂商': process_supplier(row.get('发货方', None)),
            '型号': row.get('商品名称', None) or row.get('型号', None),
            '发货数量（台）': row.get('采购数量', None) or row.get('发货数量', None),
            '收款时间': row.get('支付时间', None),
            '省市': province_city or row.get('省市', None),  # 从地址提取或使用原有字段
            '发票种类': map_invoice_type(row.get('是否需要开票', None)),
            '发票号码': row.get('发票号码', None) if map_invoice_type(row.get('是否需要开票', None)) == '开票' else None,
        }
        
        merchant_bd = row.get('商户BD', None)
        record['下订单部门'] = map_merchant_bd_to_department(merchant_bd)
        
        price = row.get('商品销售单价', None) or row.get('单价', None) or row.get('优惠价', None)
        record['单价'] = price
        record['优惠价'] = price
        
        record['实际发货台数'] = record['发货数量（台）']
        
        if record['单价'] is not None and record['发货数量（台）'] is not None:
            try:
                record['应收金额'] = float(record['单价']) * float(record['发货数量（台）'])
            except:
                record['应收金额'] = None
        else:
            record['应收金额'] = None
        
        if record['优惠价'] is not None and record['发货数量（台）'] is not None:
            try:
                record['实收金额'] = float(record['优惠价']) * float(record['发货数量（台）'])
            except:
                record['实收金额'] = 0.00
        else:
            record['实收金额'] = 0.00
        
        record['实际发货金额'] = record['实收金额']
        
        express_fee = row.get('快递费', None)
        if pd.isna(express_fee) or express_fee is None:
            record['快递费'] = 0.00
        else:
            try:
                record['快递费'] = float(express_fee)
            except:
                record['快递费'] = 0.00
        
        merchant_order_no = row.get('商城支付单号', None) or row.get('商城支付订单号', None)
        if merchant_order_no and not str(merchant_order_no).startswith('PA'):
            for col in ['商城支付单号', '商城支付订单号', '支付订单号', 'PA订单号']:
                if col in row.index:
                    val = row.get(col, None)
                    if val and str(val).startswith('PA'):
                        merchant_order_no = val
                        break
        
        record['商城支付订单号'] = merchant_order_no
        record['支付渠道'] = map_payment_channel(merchant_order_no, df_reference)
        
        if record['实际发货金额'] is not None and record['快递费'] is not None:
            try:
                record['收款金额'] = float(record['实际发货金额']) + float(record['快递费'])
            except:
                record['收款金额'] = 0.00
        else:
            record['收款金额'] = 0.00
        
        payment_status = row.get('支付状态', '')
        refund_info = process_refund_info(row, payment_status)
        record['退款-退快递费'] = refund_info['退款-退快递费']
        record['退款-退货'] = refund_info['退款-退货']
        record['退货-台数'] = refund_info['退货-台数']
        record['_需要人工校验退款'] = refund_info['需要人工校验']
        
        record['省市'] = row.get('省市', None)
        record['线上支付编号'] = row.get('渠道支付单号', None)
        record['退款单号'] = row.get('退款单号', None)
        record['线下款'] = 0.00 if pd.isna(row.get('线下款', None)) else row.get('线下款', 0.00)
        record['二次退款'] = 0.00 if pd.isna(row.get('二次退款', None)) else row.get('二次退款', 0.00)
        record['二次退款时间'] = None
        
        # 退款时间：从交易对账单中获取
        refund_time = None
        if payment_status in ['退款成功', 'REFUND_PART_SUCCESS'] and merchant_order_no and df_reference is not None:
            # 通过商城支付订单号查询交易对账单中的退款记录
            merchant_order_col = find_column(df_reference, ['商户订单号', '订单号', '商城订单号'])
            if merchant_order_col:
                merchant_order_no_str = str(merchant_order_no).strip()
                reference_df_matched = df_reference.copy()
                reference_df_matched[merchant_order_col] = reference_df_matched[merchant_order_col].astype(str).str.strip()
                
                # 查找退款订单（订单类型为退款，且退款订单原商户订单号匹配）
                if '订单类型' in df_reference.columns:
                    refund_col = find_column(df_reference, ['退款订单原商户订单号', '原商户订单号'])
                    if refund_col:
                        refund_match = reference_df_matched[
                            (reference_df_matched['订单类型'] == '退款') & 
                            (reference_df_matched[refund_col].astype(str).str.strip() == merchant_order_no_str)
                        ]
                        if not refund_match.empty:
                            transaction_time_col = find_column(df_reference, ['交易时间', '交易日期', '退款时间'])
                            if transaction_time_col:
                                refund_time_raw = refund_match.iloc[0][transaction_time_col]
                                if pd.notna(refund_time_raw):
                                    # 只保留日期部分 yyyy-mm-dd
                                    try:
                                        if isinstance(refund_time_raw, pd.Timestamp):
                                            refund_time = refund_time_raw.strftime('%Y-%m-%d')
                                        elif isinstance(refund_time_raw, str):
                                            dt = pd.to_datetime(refund_time_raw)
                                            refund_time = dt.strftime('%Y-%m-%d')
                                    except:
                                        refund_time = str(refund_time_raw).split(' ')[0] if ' ' in str(refund_time_raw) else str(refund_time_raw)
        
        record['退款时间'] = refund_time
        record['MSP单号'] = None
        record['（预留列）'] = None
        record['备注'] = None
        record['_跨月退款'] = False  # 标记是否为跨月退款
        
        result_data.append(record)
    
    # 处理交易对账单中的跨月退款订单
    if df_reference is not None and '订单类型' in df_reference.columns:
        # 查找订单类型为"退款"的记录
        refund_orders = df_reference[df_reference['订单类型'] == '退款'].copy()
        
        for idx, refund_row in refund_orders.iterrows():
            # 获取退款订单原商户订单号
            original_order_col = find_column(df_reference, ['退款订单原商户订单号', '原商户订单号', '退款原订单号'])
            if not original_order_col:
                continue
            
            original_order_no = refund_row.get(original_order_col, None)
            if pd.isna(original_order_no):
                continue
            
            # 检查订单号格式和日期（第3-10位为年月日YYYYMMDD）
            original_order_str = str(original_order_no)
            if len(original_order_str) >= 10:
                try:
                    order_date_str = original_order_str[2:10]  # 获取第3-10位
                    order_year_month = order_date_str[:6]  # 获取年月YYYYMM
                    
                    # 如果不是当月的订单
                    if order_year_month != current_month_str:
                        # 获取交易时间并格式化为yyyy-mm-dd
                        transaction_time_col = find_column(df_reference, ['交易时间', '交易日期'])
                        refund_time_raw = refund_row.get(transaction_time_col, None) if transaction_time_col else None
                        refund_time_formatted = None
                        if pd.notna(refund_time_raw):
                            try:
                                if isinstance(refund_time_raw, pd.Timestamp):
                                    refund_time_formatted = refund_time_raw.strftime('%Y-%m-%d')
                                elif isinstance(refund_time_raw, str):
                                    dt = pd.to_datetime(refund_time_raw)
                                    refund_time_formatted = dt.strftime('%Y-%m-%d')
                            except:
                                refund_time_formatted = str(refund_time_raw).split(' ')[0] if ' ' in str(refund_time_raw) else str(refund_time_raw)
                        
                        # 收款时间：从退款订单原商户订单号的第3-10位提取（YYYYMMDD）
                        payment_time_str = None
                        try:
                            # order_date_str = YYYYMMDD
                            payment_time_str = f"{order_date_str[:4]}-{order_date_str[4:6]}-{order_date_str[6:8]}"
                        except:
                            payment_time_str = None
                        
                        # ============ 跨月退款订单处理流程 ============
                        # 1. 从交易对账单获取"退款订单原商户订单号" (original_order_no)
                        # 2. 将其填入模版的"商城支付订单号"字段
                        # 3. 使用此订单号去跨月退原数据中匹配"商城支付订单号"
                        # 4. 将匹配到的订单的所有字段填充到跨月退款记录中
                        # =============================================
                        
                        # 月份格式：yyyymm-跨月退
                        cross_month_value = f"{order_year_month}-跨月退"
                        
                        # 步骤3: 在对应月份的原数据中查找匹配订单
                        # 匹配字段: 商城支付订单号 (原数据格式与当月订单相同)
                        print(f"\n发现跨月退款订单:")
                        print(f"  - 原订单月份: {order_year_month}")
                        print(f"  - 商城支付订单号(用于匹配): {original_order_no}")
                        print(f"  - 退款金额: {refund_row.get('退款金额', 0)}")
                        
                        original_order_data = None
                        if order_year_month in cross_month_dfs:
                            print(f"  - 正在 {order_year_month} 月份原数据中通过商城支付订单号匹配...")
                            original_order_data = find_cross_month_order_in_source(original_order_no, cross_month_dfs[order_year_month])
                        else:
                            print(f"  ✗ 未提供 {order_year_month} 月份的原数据文件")
                        
                        # 步骤4: 如果匹配成功，从原始订单填充所有字段到跨月退款记录
                        if original_order_data is not None:
                            print(f"  ✓ 匹配成功！正在填充完整订单信息...")
                            # 从匹配到的原始订单中提取所有字段
                            merchant_bd = original_order_data.get('商户BD', None)
                            price = original_order_data.get('商品销售单价', None) or original_order_data.get('单价', None) or original_order_data.get('优惠价', None)
                            quantity = original_order_data.get('采购数量', None) or original_order_data.get('发货数量', None)
                            express_fee = original_order_data.get('快递费', None)
                            actual_payment = original_order_data.get('实付款', None)  # 获取实付款
                            
                            # 处理金额
                            try:
                                price_float = float(price) if pd.notna(price) else 0.00
                                quantity_int = int(quantity) if pd.notna(quantity) else 0
                                express_fee_float = float(express_fee) if pd.notna(express_fee) else 0.00
                                actual_payment_float = float(actual_payment) if pd.notna(actual_payment) else 0.00
                                
                                # 从交易对账单获取退款金额（订单金额(元)字段）
                                refund_amount_field = find_column(df_reference, ['订单金额(元)', '退款金额', '订单金额'])
                                refund_amount = 0.00
                                if refund_amount_field:
                                    refund_amount = float(refund_row.get(refund_amount_field, 0)) if pd.notna(refund_row.get(refund_amount_field, 0)) else 0.00
                                
                                receivable_amount = price_float * quantity_int
                                actual_shipment_amount = receivable_amount
                                total_payment = actual_shipment_amount + express_fee_float
                                
                                # 退款信息处理（按照新规则）
                                # （1）订单金额 = 快递费：仅退快递费
                                # （2）订单金额 = 快递费 + 实付款：退快递费和退货
                                refund_express_fee = 0.00
                                refund_goods = 0.00
                                refund_quantity = 0
                                
                                if abs(refund_amount - express_fee_float) < 0.01:
                                    # 情况1：仅退快递费
                                    refund_express_fee = express_fee_float
                                    refund_goods = 0.00
                                    refund_quantity = 0
                                elif abs(refund_amount - (express_fee_float + actual_payment_float)) < 0.01:
                                    # 情况2：退快递费和退货
                                    refund_express_fee = express_fee_float
                                    refund_goods = actual_payment_float
                                    refund_quantity = quantity_int
                                else:
                                    # 其他情况：默认处理
                                    refund_express_fee = refund_amount
                                    refund_goods = 0.00
                                    refund_quantity = 0
                            except:
                                price_float = 0.00
                                quantity_int = 0
                                express_fee_float = 0.00
                                receivable_amount = 0.00
                                actual_shipment_amount = 0.00
                                total_payment = 0.00
                                refund_express_fee = 0.00
                                refund_goods = 0.00
                                refund_quantity = 0
                            
                            # 创建跨月退款记录（字段说明）
                            cross_record = {
                                # 特殊字段
                                '月份': cross_month_value,  # 格式: YYYYMM-跨月退
                                '收款时间': payment_time_str,  # 从订单号提取
                                
                                # === 以下字段从匹配到的原数据订单中填充 ===
                                '下订单部门': map_merchant_bd_to_department(merchant_bd),
                                '产品订货单号': original_order_data.get('订单号', None),
                                '商户号': original_order_data.get('买家ID', None),
                                '商户名称': original_order_data.get('买家名称', None),
                                '省市': original_order_data.get('省市', None),
                                '机具厂商': process_supplier(original_order_data.get('发货方', None)),
                                '型号': original_order_data.get('商品名称', None) or original_order_data.get('型号', None),
                                '发货数量（台）': quantity_int,
                                '单价': price_float,
                                '优惠价': price_float,
                                '应收金额': receivable_amount,
                                '实收金额': actual_shipment_amount,
                                '实际发货台数': quantity_int,
                                '实际发货金额': actual_shipment_amount,
                                '快递费': express_fee_float,
                                '收款金额': 0.00,  # 跨月退款：收款不在当月，显示为0
                                '线下款': 0.00,
                                '线上支付编号': original_order_data.get('渠道支付单号', None),
                                '发票种类': map_invoice_type(original_order_data.get('是否需要开票', None)),
                                '发票号码': original_order_data.get('发票号码', None) if map_invoice_type(original_order_data.get('是否需要开票', None)) == '开票' else None,
                                
                                # === 以下字段从交易对账单中填充 ===
                                '退款-退快递费': abs(refund_express_fee),  # 退款字段填写正数
                                '退款-退货': abs(refund_goods),  # 退款字段填写正数
                                '退货-台数': refund_quantity,  # 根据退款金额计算
                                '钱包款（实际收入金额）': 0.00 - abs(refund_express_fee) - abs(refund_goods),  # 跨月退款：收款金额为0，钱包款为负
                                '支付渠道': refund_row.get('支付渠道', None),
                                '商城支付订单号': original_order_no,  # 步骤2: 填入退款订单原商户订单号
                                '退款单号': refund_row.get('商户订单号', None),
                                '退款时间': refund_time_formatted,
                                
                                # 其他固定字段
                                '二次退款': 0.00,
                                '二次退款时间': None,
                                'MSP单号': None,
                                '（预留列）': None,
                                '备注': '跨月退款',
                                '_需要人工校验退款': False,
                                '_跨月退款': True
                            }
                        else:
                            # 如果没有找到原始订单，使用默认值
                            cross_record = {
                                '月份': cross_month_value,
                                '下订单部门': None,
                                '产品订货单号': None,
                                '商户号': None,
                                '商户名称': None,
                                '省市': None,
                                '机具厂商': None,
                                '型号': None,
                                '发货数量（台）': 0,
                                '单价': 0.00,
                                '优惠价': 0.00,
                                '应收金额': 0.00,
                                '实收金额': 0.00,
                                '收款时间': payment_time_str,
                                '实际发货台数': 0,
                                '实际发货金额': 0.00,
                                '快递费': 0.00,
                                '收款金额': 0.00,
                                '退款-退快递费': refund_row.get('退款金额', 0.00),
                                '退款-退货': 0.00,
                                '退货-台数': 0,
                                '钱包款（实际收入金额）': 0.00,
                                '线下款': 0.00,
                                '线上支付编号': None,
                                '发票种类': '不开票',
                                '发票号码': None,
                                '支付渠道': refund_row.get('支付渠道', None),
                                '商城支付订单号': original_order_no,
                                '退款单号': refund_row.get('商户订单号', None),
                                '退款时间': refund_time_formatted,
                                '二次退款': 0.00,
                                '二次退款时间': None,
                                'MSP单号': None,
                                '（预留列）': None,
                                '备注': '跨月退款-缺少原始数据',
                                '_需要人工校验退款': False,
                                '_跨月退款': True
                            }
                        
                        cross_month_refund_data.append(cross_record)
                except (ValueError, IndexError):
                    continue
    
    # 数据处理完成汇总
    print(f"\n✅ 字段映射完成")
    print(f"   - 正常订单: {len(result_data)} 条")
    print(f"   - 跨月退款订单: {len(cross_month_refund_data)} 条")
    print(f"   - 总计: {len(result_data) + len(cross_month_refund_data)} 条")
    
    # 收集跨月退款涉及的月份
    cross_month_list = set()
    for record in cross_month_refund_data:
        month_str = record.get('月份', '')
        if isinstance(month_str, str) and '-跨月退' in month_str:
            year_month = month_str.replace('-跨月退', '')
            cross_month_list.add(year_month)
            
            # 检查备注字段，如果包含"缺少原始数据"，提示用户
            if '缺少原始数据' in record.get('备注', ''):
                print(f"\n⚠ 警告: 发现 {year_month} 月份的跨月退款订单，但未提供对应月份的原数据文件")
                print(f"  - 商城支付订单号: {record.get('商城支付订单号', 'N/A')}")
                print(f"  - 说明: 跨月退款订单将根据商城支付订单号在对应月份的原数据表中匹配")
                print(f"  - 建议: 重新运行脚本时，通过 --cross-month {year_month}=对应月份文件路径 参数提供原数据")
    
    # 合并正常订单和跨月退款订单（跨月退款放在最后，中间间隔两行）
    if cross_month_refund_data:
        # 添加两个完全空的行作为间隔（不填充任何列）
        result_data.append({'_is_empty_row': True})
        result_data.append({'_is_empty_row': True})
        
        # 添加跨月退款订单
        result_data.extend(cross_month_refund_data)
    
    df_result = pd.DataFrame(result_data)
    
    # 标记空行
    if '_is_empty_row' in df_result.columns:
        empty_row_mask = df_result['_is_empty_row'].fillna(False)
    else:
        empty_row_mask = pd.Series([False] * len(df_result), index=df_result.index)
    
    for col in STATEMENT_COLUMNS:
        if col not in df_result.columns:
            df_result[col] = None
    
    # 保留必要的列（包括标记列）
    columns_to_keep = STATEMENT_COLUMNS + ['_需要人工校验退款', '_跨月退款', '_is_empty_row']
    available_columns = [col for col in columns_to_keep if col in df_result.columns]
    df_result = df_result[available_columns]
    
    # 清理数据并处理空值金额
    amount_columns = ['单价', '优惠价', '应收金额', '实收金额', '实际发货金额', '快递费', 
                      '收款金额', '退款-退快递费', '退款-退货', '钱包款（实际收入金额）', 
                      '线下款', '二次退款']
    
    for col in df_result.columns:
        if col in STATEMENT_COLUMNS:  # 只处理实际的数据列
            for idx in df_result.index:
                # 跳过空行
                if empty_row_mask.loc[idx]:
                    df_result.at[idx, col] = None
                    continue
                    
                value = df_result.at[idx, col]
                if isinstance(value, pd.Series):
                    df_result.at[idx, col] = value.iloc[0] if len(value) > 0 else None
                elif pd.isna(value):
                    # 如果是金额列，空值填入0.00
                    if col in amount_columns:
                        df_result.at[idx, col] = 0.00
                    else:
                        df_result.at[idx, col] = None
                elif isinstance(value, pd.Timestamp):
                    df_result.at[idx, col] = value.to_pydatetime()
    
    # 生成输出文件名
    if not output_path:
        today = datetime.now().strftime('%Y%m%d')
        output_path = f"{month}月对账单-盛意旺采购商城-{today}.xlsx"
    
    print(f"\n📊 生成Excel文件...")
    print(f"   - 输出路径: {output_path}")
    print(f"   - 数据行数: {len(df_result)} 行 (含空行)")
    
    # 创建Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "对账单"
    
    if 'Normal' not in wb.named_styles:
        normal_style = NamedStyle(name='Normal')
        wb.add_named_style(normal_style)
    
    ws.append(STATEMENT_COLUMNS)
    
    # 表头格式：黑色加粗文字，无背景色
    header_font = Font(name='等线', bold=True, color='000000', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment
    
    manual_check_rows = []
    cross_month_refund_rows = []  # 记录跨月退款行
    empty_rows = []  # 记录空行
    
    for idx, row in df_result.iterrows():
        # 检查是否是空行
        is_empty_row = row.get('_is_empty_row', False)
        # 处理nan值：nan应该被视为False
        if pd.isna(is_empty_row):
            is_empty_row = False
        
        if is_empty_row:
            # 空行：添加完全空的行（所有列都为None）
            row_num = ws.max_row + 1
            ws.append([None] * len(STATEMENT_COLUMNS))
            empty_rows.append(row_num)  # 记录空行行号
            continue
        
        row_data = []
        needs_check = False
        is_cross_month_refund = False
        
        for col in STATEMENT_COLUMNS:
            value = row.get(col, None)
            
            if col == '退款-退快递费' and row.get('_需要人工校验退款', False):
                needs_check = True
            
            if col == '月份' and row.get('_跨月退款', False):
                is_cross_month_refund = True
            
            if hasattr(value, '__iter__') and not isinstance(value, (str, bytes)):
                try:
                    if hasattr(value, 'iloc'):
                        value = value.iloc[0] if len(value) > 0 else None
                    elif hasattr(value, '__len__') and len(value) > 0:
                        value = value[0] if isinstance(value, (list, tuple)) else None
                    else:
                        value = None
                except:
                    value = None
            
            if pd.isna(value):
                value = None
            
            if isinstance(value, pd.Timestamp):
                value = value.to_pydatetime()
            
            row_data.append(value)
        
        row_num = ws.max_row + 1
        ws.append(row_data)
        
        # 设置全表字体为等线
        for cell in ws[row_num]:
            cell.font = Font(name='等线', size=11)
        
        if needs_check:
            manual_check_rows.append(row_num)
        
        if is_cross_month_refund:
            cross_month_refund_rows.append(row_num)
    
    # 设置公式（跳过空行和跨月退款行）
    for row in range(2, ws.max_row + 1):
        # 跳过空行
        if row in empty_rows:
            continue
        
        # 跨月退款行：收款金额(R列)不设置公式（已经填充为0.00），其他公式正常设置
        if row in cross_month_refund_rows:
            ws[f'L{row}'] = f'=J{row}*I{row}'
            ws[f'M{row}'] = f'=K{row}*I{row}'
            # R列不设置公式，保持原值0.00
            ws[f'V{row}'] = f'=R{row}-IF(ISBLANK(S{row}),0,S{row})-IF(ISBLANK(T{row}),0,T{row})'
        else:
            # 普通订单：所有公式都设置
            ws[f'L{row}'] = f'=J{row}*I{row}'
            ws[f'M{row}'] = f'=K{row}*I{row}'
            ws[f'R{row}'] = f'=P{row}+Q{row}'
            ws[f'V{row}'] = f'=R{row}-IF(ISBLANK(S{row}),0,S{row})-IF(ISBLANK(T{row}),0,T{row})'
    
    # 标注需要人工校验的退款
    if manual_check_rows:
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        for row_num in manual_check_rows:
            for col in ['S', 'T', 'U']:
                cell = ws[f'{col}{row_num}']
                cell.fill = red_fill
                cell.font = Font(name='等线', size=11)
    
    # 标注跨月退款订单（整行用#DDEBF7背景色，特定列文字为红色）
    if cross_month_refund_rows:
        cross_month_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        red_font = Font(name='等线', size=11, color='FF0000')  # 红色字体
        normal_font = Font(name='等线', size=11)
        
        for row_num in cross_month_refund_rows:
            for col_idx, cell in enumerate(ws[row_num], start=1):
                cell.fill = cross_month_fill
                # A列=1(月份), N列=14(收款时间), R列=18(收款金额), AD列=30(退款时间)
                # 这些列的文字为红色
                if col_idx in [1, 14, 18, 30]:
                    cell.font = red_font
                else:
                    cell.font = normal_font
                
                # V列（钱包款，第22列）：设置红色带括号格式
                if col_idx == 22:  # V列 = 22
                    # 使用会计格式，负数显示为红色括号（公式会自动计算）
                    cell.number_format = '#,##0.00;[Red](#,##0.00)'
    
    # 设置数字格式（在添加统计行之前，跳过空行）
    amount_cols = ['J', 'K', 'L', 'M', 'P', 'Q', 'R', 'S', 'T', 'V', 'W', 'AE']
    for col in amount_cols:
        for row in range(2, ws.max_row + 1):
            if row in empty_rows:  # 跳过空行
                continue
            cell = ws[f'{col}{row}']
            
            # V列特殊处理：跨月退款行已经设置了红色括号格式，不覆盖
            if col == 'V' and row in cross_month_refund_rows:
                continue
            
            cell.number_format = '#,##0.00'  # 所有金额列都设置格式，包括空值
            if cell.font.name != '等线':  # 如果字体未设置，设置为等线
                cell.font = Font(name='等线', size=11)
    
    quantity_cols = ['I', 'O', 'U']
    for col in quantity_cols:
        for row in range(2, ws.max_row + 1):
            if row in empty_rows:  # 跳过空行
                continue
            cell = ws[f'{col}{row}']
            if cell.value is not None:
                cell.number_format = '0'
            if cell.font.name != '等线':
                cell.font = Font(name='等线', size=11)
    
    for row in range(2, ws.max_row + 1):
        if row in empty_rows:  # 跳过空行
            continue
        cell = ws[f'A{row}']
        if cell.value is not None:
            cell.number_format = '000000'
    
    date_cols = ['N', 'AD', 'AF']
    for col in date_cols:
        for row in range(2, ws.max_row + 1):
            if row in empty_rows:  # 跳过空行
                continue
            cell = ws[f'{col}{row}']
            if cell.value is not None:
                if isinstance(cell.value, str):
                    try:
                        dt = pd.to_datetime(cell.value)
                        cell.value = dt
                    except:
                        pass
                cell.number_format = 'yyyy-mm-dd'
    
    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width
    
    # 添加统计行（在所有格式设置之后）
    summary_row_num = ws.max_row + 1
    ws.append([None] * len(STATEMENT_COLUMNS))
    
    # R列：收款金额合计
    ws[f'R{summary_row_num}'] = f'=SUM(R2:R{summary_row_num - 1})'
    # T列：退款-退快递费+退款-退货合计
    ws[f'T{summary_row_num}'] = f'=SUM(S2:S{summary_row_num - 1})+SUM(T2:T{summary_row_num - 1})'
    # U列：不显示统计数值（留空）
    # V列：钱包款（实际收入金额）合计
    ws[f'V{summary_row_num}'] = f'=SUM(V2:V{summary_row_num - 1})'
    
    # 统计行格式：微软雅黑，16号，加粗
    summary_font = Font(name='微软雅黑', size=16, bold=True)
    for col in ['R', 'T', 'V']:
        cell = ws[f'{col}{summary_row_num}']
        cell.font = summary_font
        cell.number_format = '#,##0.00'
    
    ws.freeze_panes = 'A2'
    
    wb.save(output_path)
    return output_path

if __name__ == '__main__':
    if len(sys.argv) < 4:
        print("使用方法: python generate_statement.py <原数据文件> <交易对账单文件> <月份> [输出文件]")
        print("\n可选参数:")
        print("  --cross-month YYYYMM=文件路径  指定跨月退款对应月份的原数据文件")
        print("\n示例:")
        print("  python generate_statement.py data.xlsx reference.xlsx 1")
        print("  python generate_statement.py data.xlsx reference.xlsx 1 --cross-month 202412=data_202412.xlsx")
        sys.exit(1)
    
    source_file = sys.argv[1]
    reference_file = sys.argv[2]
    month = int(sys.argv[3])
    output_path = None
    cross_month_files = {}
    
    # 解析可选参数
    i = 4
    while i < len(sys.argv):
        arg = sys.argv[i]
        if arg == '--cross-month' and i + 1 < len(sys.argv):
            # 格式：YYYYMM=文件路径
            mapping = sys.argv[i + 1]
            if '=' in mapping:
                year_month, file_path = mapping.split('=', 1)
                cross_month_files[year_month] = file_path
                print(f"已配置跨月退款数据: {year_month} -> {file_path}")
            i += 2
        else:
            # 可能是输出文件路径
            if output_path is None and not arg.startswith('--'):
                output_path = arg
            i += 1
    
    # 首次生成对账单
    print("\n开始生成对账单...")
    output = generate_statement(source_file, reference_file, month, output_path, cross_month_files)
    print(f"\n对账单已生成: {output}")
    
    # 如果输出中有警告信息，说明有跨月退款但缺少原始数据
    # 提示用户可以重新运行脚本并提供对应文件
